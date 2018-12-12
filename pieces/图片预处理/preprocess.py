import pandas as pd
import openpyxl as xl
import re
import os

file_path = './ExportResult.xlsx'
save_path = './ExportResult_update.xlsx'
def fix_hyperlink():
    wb = xl.load_workbook(file_path)
    sheet = wb['Sheet1']
    # print(sheet.cell(row = 1,column = 20)) #index from 1
    # print(sheet.cell(row=2, column=20).value)
    row_end = sheet.max_row + 1
    pattern = re.compile(r'.*"person/(.*)","(.*?)"')
    pattern1 = re.compile(r'.*"person\\(.*)","(.*?)"')
    sheet.cell(row = 1,column = 21).value = '文件名'
    for row in range(2,row_end):
        s = sheet.cell(row=row, column=20).value
        if pattern.match(s):
            filename = pattern.match(s).group(1)
        elif pattern1.match(s):
            filename = pattern1.match(s).group(1)
        sheet.cell(row = row,column = 21).value = filename
    wb.save(file_path)

# fix_hyperlink() # fixed

df = pd.read_excel(file_path)
keys = df.keys()
keys = list(keys)
keys_process = keys[:-3]
print(keys_process)
dicts = [0] * len(keys_process)

for idx in range(len(keys_process)):
    dicts[idx] = dict()

def get_tag(s,d):
    if s in d:
        return d[s]
    else:
        d[s] = len(d)
        return d[s]


filename2label = {}
for index,row in df.iterrows():
    label = ''
    for idx in range(len(keys_process)):
        s = row[keys_process[idx]]
        label += str(get_tag(s,dicts[idx]))
    row['label'] = label
    filename2label[row['文件名']] = label
df.to_excel(save_path,index = False)




def rename(path,pic_cnt):
    path = os.path.abspath(path)
    for filename in os.listdir(path):
        if filename in filename2label:
            cmd = 'mv ' + path + '/' + filename + ' ' + path + '/' + filename2label[filename] + '_' + str(pic_cnt) + '.jpg'
            os.system(cmd)
            pic_cnt += 1
    return pic_cnt
def main():
    pic_cnt = 0
    pic_cnt = rename('./raw/test/',pic_cnt)
    pic_cnt = rename('./raw/train/',pic_cnt)
    print('pic cnt: ' + str(pic_cnt))
    for idx in range(len(keys_process)):
        print('this tag: ' + keys_process[idx])
        print('the rule of labeling')
        for k,v in dicts[idx].items():
            print(k + '\t' + str(v))
# main()

def find_the_left():
    all_filename = set(filename2label.keys())
    exist = set()
    strange = set()
    path = './raw/test/'
    for filename in os.listdir(path):
        if 'person' in filename:
            strange.add(filename)

    path = './raw/train/'
    for filename in os.listdir(path):
        if 'person' in filename:
            strange.add(filename)
    diff = strange - all_filename
    print(len(diff))
    for item in diff:
        print(item)

# find_the_left()


def delete_duplicate():
    path = './raw/test/'
    path = os.path.abspath(path)
    for filename in os.listdir(path):
        if 'person' in filename:
            os.system('rm ' + path + '/' + filename)

    path = './raw/train/'
    path = os.path.abspath(path)
    for filename in os.listdir(path):
        if 'person' in filename:
            os.system('rm ' + path + '/' + filename)
delete_duplicate()

